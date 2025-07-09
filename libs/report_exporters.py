#!/usr/bin/env python3
"""
レポートエクスポーター
PDF, Excel, Markdown などの形式でレポートを出力
"""

import json
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional
from pathlib import Path
import sys
import io
import base64

# プロジェクトルートをパスに追加
sys.path.insert(0, str(Path(__file__).parent.parent))

# ロギング設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# サードパーティライブラリのモックインポート
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logger.warning("ReportLab not installed. PDF export will use fallback method.")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("OpenPyXL not installed. Excel export will use fallback method.")

class PDFExporter:
    """PDF形式でのレポートエクスポート"""
    
    def export(self, html_content: str, data: Dict[str, Any], 
              report_id: str, output_dir: Path) -> Path:
        """PDFエクスポート"""
        file_path = output_dir / f"{report_id}.pdf"
        
        if HAS_REPORTLAB:
            return self._export_with_reportlab(data, file_path)
        else:
            return self._export_fallback(html_content, data, file_path)
    
    def _export_with_reportlab(self, data: Dict[str, Any], file_path: Path) -> Path:
        """ReportLabを使用したPDF生成"""
        # PDFドキュメント作成
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # スタイル取得
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # ストーリー（コンテンツ）作成
        story = []
        
        # タイトル
        title = Paragraph(data.get('title', 'Report'), title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # 生成日時
        date_para = Paragraph(
            f"生成日時: {data.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}",
            normal_style
        )
        story.append(date_para)
        story.append(Spacer(1, 12))
        
        # セクション処理
        self._add_sections_to_story(data, story, styles)
        
        # PDF生成
        doc.build(story)
        
        return file_path
    
    def _add_sections_to_story(self, data: Dict[str, Any], story: List, styles):
        """セクションをストーリーに追加"""
        # ヘルススコア
        if 'health_score' in data:
            story.append(Paragraph("システムヘルススコア", styles['Heading2']))
            story.append(Paragraph(f"{data['health_score']}%", styles['Normal']))
            story.append(Spacer(1, 12))
        
        # 4賢者パフォーマンス
        if 'sages' in data:
            story.append(Paragraph("4賢者パフォーマンス", styles['Heading2']))
            
            # テーブルデータ準備
            table_data = [['賢者', '応答時間(ms)', '精度(%)', '可用性(%)']]
            for sage in data['sages']:
                table_data.append([
                    sage.get('name', ''),
                    str(sage.get('response_time', '')),
                    str(sage.get('accuracy', '')),
                    str(sage.get('availability', ''))
                ])
            
            # テーブル作成
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 12))
        
        # インシデント
        if 'incidents' in data and data['incidents']:
            story.append(Paragraph("インシデント", styles['Heading2']))
            for incident in data['incidents']:
                incident_text = f"<b>{incident.get('title', 'Unknown')}</b><br/>"
                incident_text += f"発生時刻: {incident.get('time', 'Unknown')}<br/>"
                incident_text += f"ステータス: {incident.get('status', 'Unknown')}"
                story.append(Paragraph(incident_text, styles['Normal']))
                story.append(Spacer(1, 6))
        
        # タスク
        if 'tasks' in data:
            story.append(Paragraph("タスク完了状況", styles['Heading2']))
            tasks = data['tasks']
            task_text = f"完了: {tasks.get('completed', 0)}<br/>"
            task_text += f"進行中: {tasks.get('in_progress', 0)}<br/>"
            task_text += f"完了率: {tasks.get('completion_rate', 0)}%"
            story.append(Paragraph(task_text, styles['Normal']))
    
    def _export_fallback(self, html_content: str, data: Dict[str, Any], 
                        file_path: Path) -> Path:
        """フォールバックPDF生成（HTMLをテキストファイルとして保存）"""
        # HTMLからテキスト抽出（簡易版）
        import re
        text_content = re.sub('<[^<]+?>', '', html_content)
        
        # メタデータ追加
        pdf_content = f"""
=== {data.get('title', 'Report')} ===
生成日時: {data.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}

{text_content}

=== レポート終了 ===
"""
        
        # テキストファイルとして保存（.pdf拡張子）
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(pdf_content)
        
        logger.warning("ReportLab not available. Generated text-based PDF.")
        return file_path

class ExcelExporter:
    """Excel形式でのレポートエクスポート"""
    
    def export(self, html_content: str, data: Dict[str, Any], 
              report_id: str, output_dir: Path) -> Path:
        """Excelエクスポート"""
        file_path = output_dir / f"{report_id}.xlsx"
        
        if HAS_OPENPYXL:
            return self._export_with_openpyxl(data, file_path)
        else:
            return self._export_fallback(data, file_path)
    
    def _export_with_openpyxl(self, data: Dict[str, Any], file_path: Path) -> Path:
        """OpenPyXLを使用したExcel生成"""
        # ワークブック作成
        wb = openpyxl.Workbook()
        
        # サマリーシート
        ws = wb.active
        ws.title = "サマリー"
        
        # タイトル
        ws['A1'] = data.get('title', 'Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"生成日時: {data.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}"
        
        # ヘルススコア
        if 'health_score' in data:
            ws['A4'] = "システムヘルススコア"
            ws['B4'] = f"{data['health_score']}%"
            ws['A4'].font = Font(bold=True)
        
        # 4賢者パフォーマンスシート
        if 'sages' in data:
            sage_ws = wb.create_sheet("4賢者パフォーマンス")
            
            # ヘッダー
            headers = ['賢者', '応答時間(ms)', '精度(%)', '可用性(%)']
            for col, header in enumerate(headers, 1):
                cell = sage_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # データ
            for row, sage in enumerate(data['sages'], 2):
                sage_ws.cell(row=row, column=1, value=sage.get('name', ''))
                sage_ws.cell(row=row, column=2, value=sage.get('response_time', 0))
                sage_ws.cell(row=row, column=3, value=sage.get('accuracy', 0))
                sage_ws.cell(row=row, column=4, value=sage.get('availability', 0))
            
            # 列幅調整
            for col in range(1, 5):
                sage_ws.column_dimensions[get_column_letter(col)].width = 15
        
        # インシデントシート
        if 'incidents' in data and data['incidents']:
            inc_ws = wb.create_sheet("インシデント")
            
            # ヘッダー
            headers = ['タイトル', '発生時刻', 'ステータス', '重要度']
            for col, header in enumerate(headers, 1):
                cell = inc_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # データ
            for row, incident in enumerate(data['incidents'], 2):
                inc_ws.cell(row=row, column=1, value=incident.get('title', ''))
                inc_ws.cell(row=row, column=2, value=incident.get('time', ''))
                inc_ws.cell(row=row, column=3, value=incident.get('status', ''))
                inc_ws.cell(row=row, column=4, value=incident.get('severity', ''))
        
        # タスクシート
        if 'tasks' in data:
            task_ws = wb.create_sheet("タスク状況")
            
            task_ws['A1'] = "タスク完了状況"
            task_ws['A1'].font = Font(size=14, bold=True)
            
            task_ws['A3'] = "完了タスク"
            task_ws['B3'] = data['tasks'].get('completed', 0)
            
            task_ws['A4'] = "進行中タスク"
            task_ws['B4'] = data['tasks'].get('in_progress', 0)
            
            task_ws['A5'] = "完了率"
            task_ws['B5'] = f"{data['tasks'].get('completion_rate', 0)}%"
        
        # 保存
        wb.save(str(file_path))
        
        return file_path
    
    def _export_fallback(self, data: Dict[str, Any], file_path: Path) -> Path:
        """フォールバックExcel生成（CSV形式）"""
        import csv
        
        # CSVファイルとして保存
        csv_path = file_path.with_suffix('.csv')
        
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # タイトル
            writer.writerow([data.get('title', 'Report')])
            writer.writerow([f"生成日時: {data.get('generated_at', '')}"])
            writer.writerow([])
            
            # ヘルススコア
            if 'health_score' in data:
                writer.writerow(['システムヘルススコア', f"{data['health_score']}%"])
                writer.writerow([])
            
            # 4賢者パフォーマンス
            if 'sages' in data:
                writer.writerow(['4賢者パフォーマンス'])
                writer.writerow(['賢者', '応答時間(ms)', '精度(%)', '可用性(%)'])
                for sage in data['sages']:
                    writer.writerow([
                        sage.get('name', ''),
                        sage.get('response_time', ''),
                        sage.get('accuracy', ''),
                        sage.get('availability', '')
                    ])
                writer.writerow([])
            
            # タスク
            if 'tasks' in data:
                writer.writerow(['タスク完了状況'])
                writer.writerow(['完了', data['tasks'].get('completed', 0)])
                writer.writerow(['進行中', data['tasks'].get('in_progress', 0)])
                writer.writerow(['完了率', f"{data['tasks'].get('completion_rate', 0)}%"])
        
        logger.warning("OpenPyXL not available. Generated CSV file.")
        return csv_path

class MarkdownExporter:
    """Markdown形式でのレポートエクスポート"""
    
    def export(self, html_content: str, data: Dict[str, Any], 
              report_id: str, output_dir: Path) -> Path:
        """Markdownエクスポート"""
        file_path = output_dir / f"{report_id}.md"
        
        # Markdownコンテンツ生成
        md_content = self._generate_markdown(data)
        
        # ファイル保存
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        return file_path
    
    def _generate_markdown(self, data: Dict[str, Any]) -> str:
        """Markdownコンテンツ生成"""
        lines = []
        
        # タイトル
        lines.append(f"# {data.get('title', 'Report')}")
        lines.append("")
        lines.append(f"**生成日時:** {data.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}")
        lines.append("")
        
        # 期間
        if 'period' in data:
            lines.append(f"**期間:** {data['period'].get('start', '')} - {data['period'].get('end', '')}")
            lines.append("")
        
        # ヘルススコア
        if 'health_score' in data:
            lines.append("## システムヘルススコア")
            lines.append(f"**{data['health_score']}%**")
            if 'health_change' in data:
                change = data['health_change']
                sign = '+' if change > 0 else ''
                lines.append(f"前日比: {sign}{change}%")
            lines.append("")
        
        # 4賢者パフォーマンス
        if 'sages' in data:
            lines.append("## 4賢者パフォーマンス")
            lines.append("")
            lines.append("| 賢者 | 応答時間(ms) | 精度(%) | 可用性(%) |")
            lines.append("|------|-------------|---------|-----------|")
            
            for sage in data['sages']:
                lines.append(
                    f"| {sage.get('name', '')} | "
                    f"{sage.get('response_time', '')} | "
                    f"{sage.get('accuracy', '')} | "
                    f"{sage.get('availability', '')} |"
                )
            lines.append("")
        
        # インシデント
        if 'incidents' in data:
            lines.append("## インシデント")
            lines.append("")
            
            if data['incidents']:
                for incident in data['incidents']:
                    severity = incident.get('severity', 'info')
                    icon = '🔴' if severity == 'critical' else '🟡' if severity == 'warning' else '🟢'
                    
                    lines.append(f"### {icon} {incident.get('title', 'Unknown')}")
                    lines.append(f"- **発生時刻:** {incident.get('time', 'Unknown')}")
                    lines.append(f"- **ステータス:** {incident.get('status', 'Unknown')}")
                    lines.append("")
            else:
                lines.append("インシデントはありません。")
                lines.append("")
        
        # タスク
        if 'tasks' in data:
            lines.append("## タスク完了状況")
            lines.append("")
            tasks = data['tasks']
            lines.append(f"- **完了タスク:** {tasks.get('completed', 0)}")
            lines.append(f"- **進行中タスク:** {tasks.get('in_progress', 0)}")
            lines.append(f"- **完了率:** {tasks.get('completion_rate', 0)}%")
            lines.append("")
        
        # メトリクス
        if 'metrics' in data:
            lines.append("## メトリクス")
            lines.append("")
            
            for metric_name, metric_value in data['metrics'].items():
                if isinstance(metric_value, dict):
                    lines.append(f"### {metric_name}")
                    for key, value in metric_value.items():
                        lines.append(f"- **{key}:** {value}")
                else:
                    lines.append(f"- **{metric_name}:** {metric_value}")
            lines.append("")
        
        # 推奨事項
        if 'recommendations' in data:
            lines.append("## 推奨事項")
            lines.append("")
            
            for i, rec in enumerate(data['recommendations'], 1):
                priority = rec.get('priority', 'medium')
                icon = '‼️' if priority == 'high' else '⚠️' if priority == 'medium' else 'ℹ️'
                
                lines.append(f"{i}. {icon} **{rec.get('action', '')}**")
                lines.append(f"   - 対象: {rec.get('target', '')}")
                lines.append(f"   - 理由: {rec.get('reason', '')}")
                lines.append(f"   - 期待改善: {rec.get('expected_improvement', '')}")
                lines.append("")
        
        return '\n'.join(lines)

if __name__ == "__main__":
    # テスト実行
    print("=== レポートエクスポーター テスト ===")
    
    # テストデータ
    test_data = {
        "title": "日次サマリーレポート",
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "period": {
            "start": "2025-07-08",
            "end": "2025-07-09"
        },
        "health_score": 95.5,
        "health_change": 2.3,
        "sages": [
            {"name": "ナレッジ賢者", "response_time": 45, "accuracy": 98, "availability": 99.5},
            {"name": "タスク賢者", "response_time": 55, "accuracy": 96, "availability": 99.0},
            {"name": "インシデント賢者", "response_time": 30, "accuracy": 99, "availability": 99.9},
            {"name": "RAG賢者", "response_time": 65, "accuracy": 94, "availability": 98.5}
        ],
        "incidents": [
            {
                "title": "API レスポンス遅延",
                "time": "2025-07-09 10:30:00",
                "status": "解決済み",
                "severity": "warning"
            }
        ],
        "tasks": {
            "completed": 45,
            "in_progress": 12,
            "completion_rate": 78.9
        }
    }
    
    # 出力ディレクトリ
    output_dir = Path("./test_reports")
    output_dir.mkdir(exist_ok=True)
    
    # 各形式でエクスポート
    print("\n1. PDFエクスポート...")
    pdf_exporter = PDFExporter()
    pdf_path = pdf_exporter.export("<html><body>Test</body></html>", test_data, "test_pdf", output_dir)
    print(f"   生成: {pdf_path}")
    
    print("\n2. Excelエクスポート...")
    excel_exporter = ExcelExporter()
    excel_path = excel_exporter.export("<html><body>Test</body></html>", test_data, "test_excel", output_dir)
    print(f"   生成: {excel_path}")
    
    print("\n3. Markdownエクスポート...")
    md_exporter = MarkdownExporter()
    md_path = md_exporter.export("<html><body>Test</body></html>", test_data, "test_markdown", output_dir)
    print(f"   生成: {md_path}")
    
    print("\nテスト完了！")